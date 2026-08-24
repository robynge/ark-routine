#!/usr/bin/env python3
"""One-time transform: Bloomberg "Transformed Data" xlsx -> data/backfill_bloomberg/.

Source: the user's local `Transformed Data/` folder (bigproject repo root) of
per-fund Bloomberg holdings exports (ARKK/ARKQ/ARKW/ARKG/ARKF/ARKX), themselves
gap-filled from the blog archive per its Transformed回填记录.xlsx (to 2026-07-09).
That folder is NOT in this repo; this script exists as provenance for how
data/backfill_bloomberg/ was produced on 2026-08-24.

For each fund it keeps ONLY dates absent from data/consolidated/<FUND>.csv at
run time (i.e. pre-2021-05-06 history plus the two known archive gaps) and maps

    Date (MM/DD/YYYY text) -> date        (ISO)
    <filename>             -> fund
    Bloomberg Name         -> company     (verbatim, e.g. "TSLA US Equity")
    Ticker                 -> ticker
    CUSIP                  -> cusip
    Weight (fraction)      -> weight      (x100, 4dp trimmed)
    Position               -> shares_held (may be fractional on blog-derived rows)
    Market Value           -> market_value (2dp trimmed)

build_history.py unions these seed files into every rebuild, with the official
archive always winning on any (date, fund) collision.

Usage: backfill_bloomberg.py --src "<Transformed Data dir>" [--repo .]
"""
import argparse
import csv
import datetime
import glob
import os

import openpyxl

HEADER = ["date", "fund", "company", "ticker", "cusip",
          "weight", "shares_held", "market_value"]

# The source xlsx contain rows stamped with NYSE holidays (e.g. "05/25/2026",
# Memorial Day) whose content is from a different era entirely -- mis-dated
# blocks, not carry-forwards. No US market holiday is a publish day, so all of
# them are excluded by rule.
ONE_OFF_CLOSURES = {"2018-12-05", "2025-01-09"}  # Bush / Carter funerals


def easter(year):
    a = year % 19
    b, c = divmod(year, 100)
    d, e = divmod(b, 4)
    g = (8 * b + 13) // 25
    h = (19 * a + b - d - g + 15) % 30
    i, k = divmod(c, 4)
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    n, p = divmod(h + l - 7 * m + 114, 31)
    return datetime.date(year, n, p + 1)


def nth_weekday(year, month, weekday, n):
    first = datetime.date(year, month, 1)
    return first + datetime.timedelta(days=(weekday - first.weekday()) % 7 + (n - 1) * 7)


def observed(d):
    if d.weekday() == 5:
        return d - datetime.timedelta(days=1)
    if d.weekday() == 6:
        return d + datetime.timedelta(days=1)
    return d


def nyse_holidays(year):
    h = {nth_weekday(year, 1, 0, 3),                      # MLK Day
         nth_weekday(year, 2, 0, 3),                      # Washington's Birthday
         easter(year) - datetime.timedelta(days=2),       # Good Friday
         nth_weekday(year, 5, 0, 5) if nth_weekday(year, 5, 0, 5).month == 5
         else nth_weekday(year, 5, 0, 4),                 # Memorial Day (last Mon May)
         observed(datetime.date(year, 7, 4)),             # Independence Day
         nth_weekday(year, 9, 0, 1),                      # Labor Day
         nth_weekday(year, 11, 3, 4),                     # Thanksgiving
         observed(datetime.date(year, 12, 25))}           # Christmas
    jan1 = datetime.date(year, 1, 1)
    if jan1.weekday() != 5:      # Sat Jan 1 is NOT observed (Fri is prior year)
        h.add(observed(jan1))
    if year >= 2022:
        h.add(observed(datetime.date(year, 6, 19)))       # Juneteenth
    return h


def iso(v):
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v or "").strip()
    try:
        m, d, y = s.split("/")
        return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
    except ValueError:
        return None


def trim(s):
    return s.rstrip("0").rstrip(".") if "." in s else s


def fmt_weight(v):
    return trim(f"{v * 100:.4f}") if isinstance(v, (int, float)) else ""


def fmt_shares(v):
    if not isinstance(v, (int, float)):
        return ""
    return str(int(round(v))) if abs(v - round(v)) < 1e-9 else trim(f"{v:.4f}")


def fmt_mv(v):
    return trim(f"{v:.2f}") if isinstance(v, (int, float)) else ""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", required=True, help="Transformed Data dir of *_Transformed_Data.xlsx")
    ap.add_argument("--repo", default=os.path.join(os.path.dirname(__file__), ".."))
    args = ap.parse_args()

    out_dir = os.path.join(args.repo, "data", "backfill_bloomberg")
    os.makedirs(out_dir, exist_ok=True)

    for path in sorted(glob.glob(os.path.join(args.src, "*_Transformed_Data.xlsx"))):
        fund = os.path.basename(path).split("_")[0]

        have = set()
        cons = os.path.join(args.repo, "data", "consolidated", f"{fund}.csv")
        with open(cons, newline="", encoding="utf-8") as fh:
            r = csv.reader(fh)
            next(r)
            have = {row[0] for row in r}

        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows, skipped_have, bad_date = [], 0, 0
        holidays_seen = {}
        hol_cache = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            d = iso(row[0])
            if not d:
                bad_date += 1
                continue
            if d in have:
                skipped_have += 1
                continue
            dt = datetime.date.fromisoformat(d)
            if dt.year not in hol_cache:
                hol_cache[dt.year] = nyse_holidays(dt.year)
            if dt.weekday() >= 5 or dt in hol_cache[dt.year] or d in ONE_OFF_CLOSURES:
                holidays_seen[d] = holidays_seen.get(d, 0) + 1
                continue
            rows.append([
                d, fund,
                str(row[1] or "").strip(),   # Bloomberg Name -> company
                str(row[2] or "").strip(),   # Ticker
                str(row[4] or "").strip(),   # CUSIP
                fmt_weight(row[11]),
                fmt_shares(row[6]),          # Position
                fmt_mv(row[9]),              # Market Value
            ])
        wb.close()

        rows.sort(key=lambda r: (r[0], -(float(r[5]) if r[5] else 0.0)))
        with open(os.path.join(out_dir, f"{fund}.csv"), "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(HEADER)
            w.writerows(rows)

        dates = sorted({r[0] for r in rows})
        print(f"{fund}: kept {len(rows)} rows / {len(dates)} dates "
              f"({dates[0]}..{dates[-1]}) | already-in-consolidated skipped {skipped_have}"
              f"{f' | bad dates {bad_date}' if bad_date else ''}")
        if holidays_seen:
            print(f"  excluded {len(holidays_seen)} non-trading dates "
                  f"({sum(holidays_seen.values())} rows): {', '.join(sorted(holidays_seen))}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
